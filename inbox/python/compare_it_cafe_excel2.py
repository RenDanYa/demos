# -*- coding: utf-8 -*-
"""对比 IT咖啡馆 Excel (提取作品数据.xlsx) 与已生成笔记, 删除已生成笔记的行。

Excel: d:\\voice\\新建文件夹\\提取作品数据.xlsx
  - 38 列, BV 号在第 3 列 (作品id)
  - 作者列第 8 列 (IT咖啡馆 261 行 + 伊江痕 1 行)

笔记: d:\\obsidian\\视频\\resource\\IT咖啡馆-*.md
  - frontmatter 作品网址 提取 BV 号

行为:
- 已生成笔记的 BV 号 → 删除该行
- 未生成笔记的 BV 号 → 保留该行
"""
import os
import re
import shutil
from pathlib import Path

import openpyxl

EXCEL_PATH = Path(r"d:\voice\新建文件夹\提取作品数据.xlsx")
NOTE_DIR = Path(r"d:\obsidian\视频\resource")
BACKUP_PATH = EXCEL_PATH.with_suffix(".bak.xlsx")
BV_COL = 3  # 作品id 列


def extract_bv_from_notes():
    """从 IT咖啡馆-*.md 笔记中提取 BV 号集合"""
    bv_pattern = re.compile(r"BV[0-9A-Za-z]{10}")
    url_pattern = re.compile(r'^作品网址:\s*"?([^"\s]+)"?', re.MULTILINE)

    bv_set = set()
    files = sorted([f for f in os.listdir(NOTE_DIR) if f.startswith("IT咖啡馆-") and f.endswith(".md")])
    print(f"扫描笔记目录: {NOTE_DIR}")
    print(f"IT咖啡馆笔记数: {len(files)}")

    for fname in files:
        path = NOTE_DIR / fname
        try:
            content = path.read_text(encoding="utf-8")
            m = url_pattern.search(content)
            if m:
                bv_match = bv_pattern.search(m.group(1))
                if bv_match:
                    bv_set.add(bv_match.group(0))
                    continue
            bv_matches = bv_pattern.findall(content)
            if bv_matches:
                bv_set.add(bv_matches[0])
        except Exception as e:
            print(f"  读取失败 {fname}: {str(e)[:60]}")

    print(f"提取到 BV 号: {len(bv_set)} 个")
    return bv_set


def main():
    # 1. 提取笔记 BV 号
    note_bvs = extract_bv_from_notes()

    # 2. 读取 Excel
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active
    print(f"\nExcel: {EXCEL_PATH.name}")
    print(f"  Sheet: {ws.title}, 总行数: {ws.max_row} (含表头)")
    print(f"  BV 号列: Col {BV_COL} ({ws.cell(row=1, column=BV_COL).value})")

    # 3. 收集每行 BV 号
    bv_rows = {}  # {BV号: [行号]}
    for row_idx in range(2, ws.max_row + 1):
        bv = ws.cell(row=row_idx, column=BV_COL).value
        if bv:
            bv = str(bv).strip()
            bv_rows.setdefault(bv, []).append(row_idx)

    excel_bvs = set(bv_rows.keys())
    print(f"  不重复 BV 号: {len(excel_bvs)} 个")

    # 4. 对比
    generated = excel_bvs & note_bvs
    missing = excel_bvs - note_bvs

    print(f"\n对比结果:")
    print(f"  Excel 中 BV 号: {len(excel_bvs)} 个")
    print(f"  已生成笔记:    {len(generated)} 个 BV 号")
    print(f"  未生成笔记:    {len(missing)} 个 BV 号")

    generated_rows = sum(len(bv_rows[bv]) for bv in generated)
    missing_rows = sum(len(bv_rows[bv]) for bv in missing)
    print(f"  已生成笔记行数: {generated_rows} 行 (将删除)")
    print(f"  未生成笔记行数: {missing_rows} 行 (将保留)")

    if missing:
        print(f"\n未生成笔记的 BV 号 (前 20 个):")
        for bv in sorted(missing)[:20]:
            sample_row = bv_rows[bv][0]
            title = ws.cell(row=sample_row, column=10).value or ""  # 作品标题列
            print(f"  {bv}  行{bv_rows[bv]}  {title[:50]}")
        if len(missing) > 20:
            print(f"  ... 还有 {len(missing) - 20} 个")

    # 5. 备份 + 删除
    print(f"\n备份原文件 → {BACKUP_PATH.name}")
    shutil.copy2(EXCEL_PATH, BACKUP_PATH)

    rows_to_delete = sorted(
        [r for bv in generated for r in bv_rows[bv]], reverse=True
    )
    print(f"删除 {len(rows_to_delete)} 行...")
    for row_idx in rows_to_delete:
        ws.delete_rows(row_idx, 1)

    wb.save(EXCEL_PATH)
    wb.close()

    print(f"\n已保存: {EXCEL_PATH}")
    print(f"  剩余行数: {ws.max_row - 1} 行")
    print(f"  备份: {BACKUP_PATH}")
    print("完成")


if __name__ == "__main__":
    main()
