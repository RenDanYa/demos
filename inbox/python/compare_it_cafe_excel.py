# -*- coding: utf-8 -*-
"""对比 IT咖啡馆 Excel 与已生成笔记, 删除已生成笔记的行, 保留未生成的。

Excel: d:\python_code\Selenium\test\0_前置...\b站所有视频详细列表.xlsx
笔记:  d:\obsidian\视频\resource\IT咖啡馆-*.md

对比维度: BV 号 (笔记 frontmatter 作品网址 提取)
- Excel Col 1: BV号 (如 BV1nfuT6EELu)
- 笔记 frontmatter: 作品网址: "https://www.bilibili.com/video/BV1mH4y1a7mk"

行为:
- 已生成笔记的 BV 号 → 删除 Excel 中该 BV 号所有行
- 未生成笔记的 BV 号 → 保留 Excel 中该 BV 号所有行

输出: 覆盖原 Excel (保留未生成笔记的行)
"""
import os
import re
import sys
import shutil
from pathlib import Path

import openpyxl

EXCEL_PATH = Path(
    r"d:\python_code\Selenium\test\0_前置_在生成笔记前需要判断是否有分集生成完整链接\b站所有视频详细列表.xlsx"
)
NOTE_DIR = Path(r"d:\obsidian\视频\resource")
BACKUP_PATH = EXCEL_PATH.with_suffix(".bak.xlsx")


def extract_bv_from_notes():
    """从 IT咖啡馆-*.md 笔记中提取 BV 号集合

    解析 frontmatter 的 作品网址 字段。
    返回: set[str] BV 号集合
    """
    bv_pattern = re.compile(r"BV[0-9A-Za-z]{10}")
    url_pattern = re.compile(r'^作品网址:\s*"?([^"\s]+)"?', re.MULTILINE)

    bv_set = set()
    files = sorted([f for f in os.listdir(NOTE_DIR) if f.startswith("IT咖啡馆-") and f.endswith(".md")])
    print(f"扫描笔记目录: {NOTE_DIR}")
    print(f"IT咖啡馆笔记数: {len(files)}")

    for fname in files:
        path = NOTE_DIR / fname
        try:
            content = path.read_text(encoding="utf-8")
            # 优先从 frontmatter 的 作品网址 提取
            m = url_pattern.search(content)
            if m:
                url = m.group(1)
                bv_match = bv_pattern.search(url)
                if bv_match:
                    bv_set.add(bv_match.group(0))
                    continue
            # 兜底: 从正文链接提取
            bv_matches = bv_pattern.findall(content)
            if bv_matches:
                bv_set.add(bv_matches[0])
        except Exception as e:
            print(f"  读取失败 {fname}: {str(e)[:60]}")

    print(f"提取到 BV 号: {len(bv_set)} 个")
    return bv_set


def load_excel_bv_rows():
    """读取 Excel, 返回 {BV号: [行号列表]}

    行号从 2 开始 (1 是表头)
    """
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active
    print(f"\nExcel: {EXCEL_PATH.name}")
    print(f"  Sheet: {ws.title}, 总行数: {ws.max_row} (含表头)")
    print(f"  列: {[c.value for c in ws[1]]}")

    bv_rows = {}
    for row_idx in range(2, ws.max_row + 1):
        bv = ws.cell(row=row_idx, column=1).value
        if bv:
            bv = str(bv).strip()
            bv_rows.setdefault(bv, []).append(row_idx)

    print(f"  不重复 BV 号: {len(bv_rows)} 个")
    return wb, ws, bv_rows


def main():
    # 1. 提取笔记 BV 号
    note_bvs = extract_bv_from_notes()

    # 2. 读取 Excel
    wb, ws, bv_rows = load_excel_bv_rows()

    # 3. 对比
    excel_bvs = set(bv_rows.keys())
    generated = excel_bvs & note_bvs      # 已生成笔记的 BV 号
    missing = excel_bvs - note_bvs        # 未生成笔记的 BV 号

    print(f"\n对比结果:")
    print(f"  Excel 中 BV 号: {len(excel_bvs)} 个")
    print(f"  已生成笔记:    {len(generated)} 个 BV 号")
    print(f"  未生成笔记:    {len(missing)} 个 BV 号")

    # 统计已生成笔记对应的行数
    generated_rows = sum(len(bv_rows[bv]) for bv in generated)
    missing_rows = sum(len(bv_rows[bv]) for bv in missing)
    print(f"  已生成笔记行数: {generated_rows} 行 (将删除)")
    print(f"  未生成笔记行数: {missing_rows} 行 (将保留)")

    # 打印未生成笔记的 BV 号 (前 20 个)
    if missing:
        print(f"\n未生成笔记的 BV 号 (前 20 个):")
        for bv in sorted(missing)[:20]:
            sample_row = bv_rows[bv][0]
            title = ws.cell(row=sample_row, column=4).value or ""
            print(f"  {bv}  行{bv_rows[bv]}  {title[:50]}")
        if len(missing) > 20:
            print(f"  ... 还有 {len(missing) - 20} 个")

    # 4. 备份 + 删除已生成笔记的行
    print(f"\n备份原文件 → {BACKUP_PATH.name}")
    shutil.copy2(EXCEL_PATH, BACKUP_PATH)

    # 从后往前删除行 (避免行号偏移)
    rows_to_delete = sorted(
        [r for bv in generated for r in bv_rows[bv]], reverse=True
    )
    print(f"删除 {len(rows_to_delete)} 行 (已生成笔记的 BV 号对应行)...")
    for row_idx in rows_to_delete:
        ws.delete_rows(row_idx, 1)

    # 5. 保存
    wb.save(EXCEL_PATH)
    wb.close()

    print(f"\n已保存: {EXCEL_PATH}")
    print(f"  剩余行数: {ws.max_row - 1} 行 (原 {ws.max_row + len(rows_to_delete) - 1} 行)")
    print(f"  备份: {BACKUP_PATH}")
    print("完成")


if __name__ == "__main__":
    main()
